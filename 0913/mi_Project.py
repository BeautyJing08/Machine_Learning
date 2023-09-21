
import os
import openpyxl
from openpyxl.styles import Font
import datetime
import time

print("===============================================")
print("=====Start the code to modify file's name======")

folder_path = "path/to/your/000/folder"  # 請將路徑替換為你的資料夾路徑

wb = openpyxl.Workbook()  # 打開openpyxl的套件功能 #開啟一個新的Excel檔案
sheet = wb.active  # 讓這個excel打開sheet
sheet.title = "Filename Mapping"  # 命名這個sheet
sheet["A1"] = "old_name"  # A1欄名稱為old_name
sheet["B1"] = "new_name"  # B1欄名稱為new_name
sheet["C1"] = "modify time"  # C1欄名稱為modify time
sheet.column_dimensions["A"].width = 50  # 設定欄寬
sheet.column_dimensions["B"].width = 50  # 設定欄寬
sheet.column_dimensions["C"].width = 50  # 設定欄寬

current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
###################現在執行程式碼###############################
# folder_path = "C:\\Users\\CHINGWANG\\Desktop\\08000"  # 請將路徑替換為你的資料夾路徑
beforeFolderPath = os.getcwd()
print(f"beforeFolderPath = {beforeFolderPath}")
print(type(beforeFolderPath))
folder_path = beforeFolderPath

afterFolderPath = r"{folder_path}".format \
    (folder_path=folder_path)  # 你可以在 f-string 外面將 afterFolderPath 格式化，然後再將整個字串指定給 folder_path 變數。
print(f"afterFolderPath = {afterFolderPath}")
print(type(afterFolderPath))

folder_path = afterFolderPath


# folder_path = os.getcwd()
print(f"目前的路徑為：{folder_path}")


path_parts = folder_path.split("\\")  # 用反斜線來把路徑切開成一個list，放進去好幾個element
# print(type(path_parts))

micap = path_parts[-1]  # 這個list中最後一段內容即為micap
micron_locate = "F16_0A3_"  # 這個是自己輸入的機台位置名稱

print(f"micap= {micap}")

#########################################
#########################################
#####先確認是否有micap_
excel_name = f"{micap}_change_process"  # 設定檔名
excel_path = os.path.join(folder_path, f"{excel_name}.xlsx")  # 在這個資料夾當中，增加這個excel的位置


if os.path.exists(excel_path):  # 若在這整段路徑中，已經有出現相同{micap}_change_process.xlsx == 已經執行過這個程式，所以不執行以下程式
    print("已曾經執行過此程式執行檔，故忽略此次執行。")
    print("若仍要執行程式，請將micap{micap}_change_process.xlsx檔案刪除，再重新執行程式")
else:  # 沒有出現micap{micap}_change_process.xlsx代表沒有執行過程式，故進行程式碼

    # replace_dict_CompanyName = {"A1": "AU","B1": "BF","F1": "FC","G1": "GL","G2": "GI","H1": "HK","J1": "JE"}
    replace_dict_CompanyName = {
        "A1": "AU", "利亞洲": "AU",
        "B1": "BF", "鉑峰": "BF",
        "F1": "FC", "富呈": "FC",
        "G1": "GL", "綠點": "GL",
        "G2": "GI", "京英": "GI",
        "H1": "HK", "協崑": "HK",
        "J1": "JE", "仲棠": "JE",
        "J2": "JY", "哲毅": "JY",
        "J3": "JH", "金鴻興" : "JH",
        "J4": "JF", "久福" : "JF",
        "K1": "KY", "廣毅" : "KY",
        "L1": "LY", "麗邑": "LY",
        "L2": "LJ", "力捷" : "LJ",
        "M1": "MI", "帆宣": "MI",
        "O1": "OR", "統怡" : "OR",
        "P1": "PD", "派德" :  "PD",
        "P2": "PJ", "品嘉" :  "PJ",
        "P3": "YR", "耀儒" : "YR",
        "R1": "RY", "銳澤" : "RY",
        "S1": "SW", "茂迅" : "SW",
        "S2": "SP", "萱譜" : "SP",
        "T1": "TT", "信紘" : "TT",
        "T2": "TY", "大陽日酸" : "TY",
        "U1": "UY", "晃誼" : "UY",
        "U2": "UI", "漢唐" : "UI",
        "U3": "UC", "UCAN" : "UC",
        "W1": "WT", "漢科" : "WT",
        "Y1": "YC", "宜程" :  "YC",
        "Y2": "YP", "育朋" : "YP"
    }
    replace_dict_systemName = {
        "CSD": "DATM",
        "Foundation": "FUDN", "FD" : "FUDN",
        "Power": "ELEC", "PO" : "ELEC",
        "PL" : "PMPL",
        "Exh" : "EXHT",
        "EXH" : "EXHT",
        "UPW" : "WUPW",
        # "UP" :  "WUPW",
        "ROR" : "WROR",
        "DRAIN" : "DRPR", "DR" : "DRPR",
        "WPDS": "DRPR",
        "WCCS" : "DRPR",
        "PCDA" : "ACDA",
        "BG" : "BGAS",
        "NG" :"BGAS",
        "PV": "ASPV",
        "NG": "NGAS",
        "PCW": "WPCW", "PW": "WPCW",
        "JW": "JUNK",
        "SG": "SGAS",
        "Chem": "CHEM",
        "CH": "CHEM",
        "SLURRY": "CHEM",
        "Int": "RAFL",
        "IN": "RAFL",
        "GD": "LFSY",
        "I&C": "INCL",
        "LSMK": "LSMK"
    }

    for filename in os.listdir(folder_path):
        new_filename = filename  # 創建一個新變數new_filename來裝 filename
        base_name, extension = os.path.splitext(filename)  # 分離檔名和副檔名 base_name= 檔名 ， extension = 副檔名
        if extension == ".rvt":  # 要副檔名是.rvt才要更名
            if base_name.startswith(micap):
                base_name = base_name.replace(f"{micap}", f"{micron_locate}{micap}")  # 切換micap名稱

                micap_part = base_name[:13]
                print(f"micap_part =  {micap_part}")

                base_name = base_name[13:]
                print(f"base_name =  {base_name}")

                for key, value in replace_dict_CompanyName.items():
                    base_name = base_name.replace(key, value)  # 一次替換所有的公司名稱

                for key, value in replace_dict_systemName.items():
                    base_name = base_name.replace(key, value)  # 一次替換所有的系統名稱名稱
                new_filename = micap_part + base_name
                new_filepath = os.path.join(folder_path, new_filename + extension)  # 把修改好的新檔案名稱與副檔名綁在一起
                old_filepath = os.path.join(folder_path, filename)

                if new_filename.endswith("_V2") or new_filename.endswith("_v2"):
                    base_name = base_name[:-3]  # 刪除有_V2的內容
                    new_filename = micap_part + base_name
                    new_filepath = os.path.join(folder_path, new_filename + extension)  # 把修改好的新檔案名稱與副檔名綁在一起
                    old_filepath = os.path.join(folder_path, filename)

                if not os.path.exists(new_filepath):  # 檢查新的檔名是否已存在
                    os.rename(old_filepath, new_filepath)

            if not new_filename.endswith(extension):
                new_filename = new_filename + extension

            # new_filename = new_filename + extension # 最後把名字+副檔名綁在一起
            sheet.append([filename, new_filename if new_filename != filename else ""])  # 若有改名字的話 B欄要插入新欄位名稱，若無，則自行修改
            for row_index, (filename, new_filename) in enumerate(sheet.iter_rows(min_row=2, max_col=2),
                                                                 start=2):  # 加入插入時間
                if new_filename.value != filename.value:
                    sheet.cell(row=row_index, column=3, value=current_time)

        else:  # 若不是.rvt檔案的話，也會插入進excel，但是名稱不會更改
            red_font = Font(color="FF0000")
            if not new_filename.endswith(extension):
                new_filename = new_filename + extension
            # new_filename = base_name + extension # 最後把名字+副檔名綁在一起
            sheet.append([filename, new_filename if new_filename != filename else ""])  # 若有改名字的話 B欄要插入新欄位名稱，若無，則自行修改
            for row_index, (filename, new_filename) in enumerate(sheet.iter_rows(min_row=2, max_col=2),
                                                                 start=2):  # 加入插入時間
                if new_filename.value != filename.value:
                    sheet.cell(row=row_index, column=3, value=current_time).font = red_font  # 設定紅色字體

    ######################儲存excel檔案####################################
    excel_name = f"{micap}_change_process"  # 設定檔名
    excel_path = os.path.join(folder_path, f"{excel_name}.xlsx")  # 在這個資料夾當中，增加這個excel的位置
    wb.save(excel_path)  # 另存新檔

    #############完成程式，寫出順利完成######################################
    print(f"順利執行程式，Mapping data saved to {excel_path}")
    print("==================================================")

    # ##結束以上code，最後把此份Excel打開##

    # os.startfile(excel_path) ##結束以上code，最後把此份Excel打開##
    # time.sleep(5) #給時間讓Excel檔案打開

    ###用另外一種方式來開啟檔案

    import subprocess
    import shutil

    temp_excel_path = os.path.join(folder_path, f"{excel_name}_temp.xlsx")
    shutil.copyfile(excel_path, temp_excel_path)

    # 開啟副本的Excel檔案

    subprocess.Popen([temp_excel_path], shell=True)

    # 等待時間確保Excel開啟
    time.sleep(5)

    # 等待使用者關閉Excel檔案
    print("--------------說明如下------------------")
    print("請注意此處是要請您將要修改的Excel內容修改完之後直接按存檔，並且關閉檔案，後續才會自動將暫存檔案關閉")
    print("再到cmd(終端機)內點擊Enter，就會發現戰存檔被關閉")
    print("----------------------------------------")
    input("請確認你已經完成修改並關閉 Excel 檔案後，按 Enter 繼續...")
    # 關閉 Excel 檔案 進程
    # subprocess.call(["taskkill","/f","/im", "excel.exe"])

    shutil.copyfile(temp_excel_path, excel_path)
    # 刪除暫時的Exc